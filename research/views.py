from django.contrib.auth.views import LoginView as BaseLoginView
from django.contrib.auth.views import auth_logout
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse_lazy, reverse
from django.shortcuts import redirect, get_object_or_404
from django.views import generic
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


from research.forms import AuthForm, ExperimentForm
from research.models import Experiment
from users.decorators import user_has_access


class LoginView(BaseLoginView):
    form_class = AuthForm
    template_name = "login.html"
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy("research:experiment_list")


@method_decorator(user_has_access, "dispatch")
class LogoutView(generic.View):
    def get(self, request, *args, **kwargs):
        auth_logout(request)
        return HttpResponseRedirect(reverse_lazy("research:login"))


@method_decorator(user_has_access, "dispatch")
@method_decorator(never_cache, "dispatch")
class IndexView(generic.TemplateView):
    template_name = "index.html"


@method_decorator(user_has_access, "dispatch")
@method_decorator(never_cache, "dispatch")
class ExperimentCreateView(generic.FormView):
    template_name = "experiment_form.html"
    form_class = ExperimentForm
    success_url = reverse_lazy("research:experiment_list")

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            experiment = form.save()
            experiment.calculate()
            messages.success(self.request, "Эксперимент успешно проведен!")
            return redirect("research:experiment_results", pk=experiment.id)
        else:
            return self.form_invalid(form)


@method_decorator(user_has_access, "dispatch")
@method_decorator(never_cache, "dispatch")
class ExperimentResultsView(generic.DetailView):
    model = Experiment
    template_name = "experiment_detail.html"
    context_object_name = "experiment"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        experiment = self.get_object()
        if experiment.results:
            context["chart_data"] = self.prepare_chart_data(experiment.results)
        return context

    def prepare_chart_data(self, results):
        """Подготавливает данные для графиков Chart.js"""
        chart_data = {}
        if "result_t_const" in results:
            chart_data["constant_temp"] = {
                "labels": list(results["result_t_const"].keys()),
            }
            data_tmin = []
            data_tmax = []
            data_tavg = []
            for obj in results["result_t_const"].values():
                data_tmin.append(obj["tmin_const"])
                data_tmax.append(obj["tmax_const"])
                data_tavg.append(obj["tavg_const"])

            chart_data["constant_temp"]["datasets"] = [
                {
                    "label": f"Tемпература = {self.object.t_min}°C",
                    "data": data_tmin,
                    "borderColor": "#ff6384",
                    "backgroundColor": "rgba(255, 99, 132, 0.1)",
                    "tension": 0.4,
                },
                {
                    "label": f"Tемпература = {self.object.t_max}°C",
                    "data": data_tmax,
                    "borderColor": "#36a2eb",
                    "backgroundColor": "rgba(54, 162, 235, 0.1)",
                    "tension": 0.4,
                },
                {
                    "label": f"Tемпература = {(self.object.t_min + self.object.t_max) / 2}°C",
                    "data": data_tavg,
                    "borderColor": "#4bc0c0",
                    "backgroundColor": "rgba(75, 192, 192, 0.1)",
                    "tension": 0.4,
                },
            ]

        if "result_tau_const" in results:
            chart_data["constant_time"] = {
                "labels": list(results["result_tau_const"].keys())
            }
            data_taumin = []
            data_taumax = []
            data_tauavg = []
            for obj in results["result_tau_const"].values():
                data_taumin.append(obj["taumin_const"])
                data_taumax.append(obj["taumax_const"])
                data_tauavg.append(obj["tauavg_const"])

            chart_data["constant_time"]["datasets"] = [
                {
                    "label": f"Время = {self.object.tau_min} мин",
                    "data": data_taumin,
                    "borderColor": "#ff9f40",
                    "backgroundColor": "rgba(255, 159, 64, 0.1)",
                    "tension": 0.4,
                },
                {
                    "label": f"Время = {self.object.tau_max} мин",
                    "data": data_taumax,
                    "borderColor": "#9966ff",
                    "backgroundColor": "rgba(153, 102, 255, 0.1)",
                    "tension": 0.4,
                },
                {
                    "label": f"Время = {(self.object.tau_min + self.object.tau_max) / 2} мин",
                    "data": data_tauavg,
                    "borderColor": "#ffcd56",
                    "backgroundColor": "rgba(255, 205, 86, 0.1)",
                    "tension": 0.4,
                },
            ]

        return chart_data


@method_decorator(user_has_access, "dispatch")
@method_decorator(never_cache, "dispatch")
class ExperimentListView(generic.ListView):
    model = Experiment
    template_name = "experiment_list.html"
    context_object_name = "experiments"
    ordering = ["-created_at"]
    paginate_by = 10


@method_decorator(user_has_access, "dispatch")
@method_decorator(never_cache, "dispatch")
class ExperimentRecalculateView(generic.View):
    def post(self, request, pk):
        experiment = get_object_or_404(Experiment, pk=pk)

        try:
            experiment.calculate()
            experiment.save()
            messages.success(request, "Результаты эксперимента успешно пересчитаны!")
        except Exception as e:
            messages.error(request, f"Ошибка при пересчете: {str(e)}")

        return redirect("research:experiment_results", pk=experiment.id)


def export_experiment_to_excel(request, pk):
    experiment = get_object_or_404(Experiment, pk=pk)

    wb = openpyxl.Workbook()

    ws_main = wb.active
    ws_main.title = "Информация об эксперименте"

    header_font = Font(size=14, bold=True, color="FFFFFF")
    subheader_font = Font(size=12, bold=True)
    normal_font = Font(size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="2c3e50", end_color="2c3e50", fill_type="solid"
    )
    subheader_fill = PatternFill(
        start_color="3498db", end_color="3498db", fill_type="solid"
    )

    ws_main.merge_cells("A1:E1")
    title_cell = ws_main["A1"]
    title_cell.value = f"Результаты эксперимента №{experiment.id}"
    title_cell.font = header_font
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.fill = header_fill

    info_data = [
        (
            "Математическая модель:",
            experiment.math_model.name if experiment.math_model else "Не указана",
        ),
        ("Дата проведения:", experiment.created_at.strftime("%d.%m.%Y %H:%M")),
        ("Диапазон температуры:", f"{experiment.t_min}°C - {experiment.t_max}°C"),
        ("Шаг температуры:", f"{experiment.delta_t}°C"),
        ("Средняя температура:", f"{experiment.t_avg}°C"),
        ("Диапазон времени:", f"{experiment.tau_min} - {experiment.tau_max} мин"),
        ("Шаг времени:", f"{experiment.delta_tau} мин"),
        ("Среднее время:", f"{experiment.tau_avg} мин"),
    ]

    for i, (label, value) in enumerate(info_data, start=3):
        ws_main[f"A{i}"] = label
        ws_main[f"A{i}"].font = subheader_font
        ws_main[f"B{i}"] = value
        ws_main[f"B{i}"].font = normal_font

    ws_main["A12"] = "Коэффициенты математической модели:"
    ws_main["A12"].font = Font(size=12, bold=True)

    coefficients = [
        ("a0", experiment.math_model.a_0),
        ("a1", experiment.math_model.a_1),
        ("a2", experiment.math_model.a_2),
        ("a3", experiment.math_model.a_3),
        ("a4", experiment.math_model.a_4),
        ("a5", experiment.math_model.a_5),
        ("a6", experiment.math_model.a_6),
        ("a7", experiment.math_model.a_7),
        ("a8", experiment.math_model.a_8),
    ]

    for i, (coef_name, coef_value) in enumerate(coefficients, start=13):
        ws_main[f"A{i}"] = f"{coef_name}:"
        ws_main[f"A{i}"].font = subheader_font
        ws_main[f"B{i}"] = coef_value
        ws_main[f"B{i}"].font = normal_font

    ws_main[f"A{i + 1}"] = "Формула рассчета"
    ws_main[f"A{i + 1}"].font = Font(size=12, bold=True)
    ws_main[f"A{i + 2}"] = (
        "y = a₀ + a₁·t + a₂·τ + a₃·t·τ + a₄·t² + a₅·τ² + a₆·t²·τ + a₇·t·τ² + a₈·t²·τ²"
    )
    ws_main[f"A{i + 2}"].font = Font(size=12, bold=True)

    if experiment.results and "result_t_const" in experiment.results:
        ws_temp = wb.create_sheet("При постоянной температуре")

        ws_temp.merge_cells("A1:D1")
        title_cell = ws_temp["A1"]
        title_cell.value = "Результаты при постоянной температуре"
        title_cell.font = header_font
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = header_fill

        ws_temp["A2"] = "Время (мин)"
        ws_temp["B2"] = f"T = {experiment.t_min}°C"
        ws_temp["C2"] = f"T = {experiment.t_max}°C"
        ws_temp["D2"] = f"T = {experiment.t_avg}°C"

        for col in ["A", "B", "C", "D"]:
            cell = ws_temp[f"{col}2"]
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal="center")
            cell.fill = subheader_fill
            cell.border = border

        row = 3
        result_t_const = experiment.results["result_t_const"]
        sorted_times = sorted(result_t_const.keys())

        for time in sorted_times:
            data = result_t_const[time]
            ws_temp[f"A{row}"] = time
            ws_temp[f"B{row}"] = data.get("tmin_const", "")
            ws_temp[f"C{row}"] = data.get("tmax_const", "")
            ws_temp[f"D{row}"] = data.get("tavg_const", "")

            for col in ["A", "B", "C", "D"]:
                cell = ws_temp[f"{col}{row}"]
                cell.font = normal_font
                cell.border = border
                if col != "A":
                    cell.alignment = Alignment(horizontal="center")

            row += 1

        ws_temp[f"A{row}"] = f"Всего записей: {len(sorted_times)}"
        ws_temp[f"A{row}"].font = Font(bold=True)
        ws_temp.merge_cells(f"A{row}:D{row}")

    if experiment.results and "result_tau_const" in experiment.results:
        ws_time = wb.create_sheet("При постоянном времени")

        ws_time.merge_cells("A1:D1")
        title_cell = ws_time["A1"]
        title_cell.value = "Результаты при постоянном времени"
        title_cell.font = header_font
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = header_fill

        ws_time["A2"] = "Температура (°C)"
        ws_time["B2"] = f"τ = {experiment.tau_min} мин"
        ws_time["C2"] = f"τ = {experiment.tau_max} мин"
        ws_time["D2"] = f"τ = {experiment.tau_avg} мин"

        for col in ["A", "B", "C", "D"]:
            cell = ws_time[f"{col}2"]
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal="center")
            cell.fill = subheader_fill
            cell.border = border

        row = 3
        result_tau_const = experiment.results["result_tau_const"]
        sorted_temps = sorted(result_tau_const.keys())

        for temp in sorted_temps:
            data = result_tau_const[temp]
            ws_time[f"A{row}"] = temp
            ws_time[f"B{row}"] = data.get("taumin_const", "")
            ws_time[f"C{row}"] = data.get("taumax_const", "")
            ws_time[f"D{row}"] = data.get("tauavg_const", "")

            for col in ["A", "B", "C", "D"]:
                cell = ws_time[f"{col}{row}"]
                cell.font = normal_font
                cell.border = border
                if col != "A":
                    cell.alignment = Alignment(horizontal="center")

            row += 1

        ws_time[f"A{row}"] = f"Всего записей: {len(sorted_temps)}"
        ws_time[f"A{row}"].font = Font(bold=True)
        ws_time.merge_cells(f"A{row}:D{row}")

    # Настройка ширины колонок для всех листов
    for ws in wb.worksheets:
        if ws.title == "Информация об эксперименте":
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 20
        elif ws.title in ["При постоянной температуре", "При постоянном времени"]:
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
        elif ws.title == "Формула расчета":
            ws.column_dimensions["A"].width = 50

    # Создаем HTTP response с файлом Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="experiment_{experiment.id}_results.xlsx"'
    )

    wb.save(response)
    return response
